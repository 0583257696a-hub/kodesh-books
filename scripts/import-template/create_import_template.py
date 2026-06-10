from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

out_dir = Path("public/downloads")
out_dir.mkdir(parents=True, exist_ok=True)
out_file = out_dir / "otzar_hakodesh_books_import_template.xlsx"

categories = [
    "חומשים",
    "גמרות ומשניות",
    "הלכה",
    "חסידות וקבלה",
    "ספרי ילדים ונוער",
    "מוסר ומחשבה",
    "ספרי הרב עובדיה יוסף",
    'ספרי חב"ד',
]

headers = [
    "SKU",
    "Barcode",
    "BookName",
    "SubCategory",
    "AdditionalCategories",
    "Author",
    "Rabbi",
    "Publisher",
    "Description",
    "LongDescription",
    "Price",
    "SalePrice",
    "StockQuantity",
    "Weight",
    "Language",
    "ImageFileName",
    "Featured",
    "Recommended",
    "New",
    "Active",
]

hebrew_labels = [
    'מק"ט',
    "ברקוד",
    "שם הספר",
    "תת קטגוריה",
    "קטגוריות נוספות",
    "מחבר",
    "רב",
    "הוצאה לאור",
    "תיאור קצר",
    "תיאור ארוך",
    "מחיר",
    "מחיר מבצע",
    "כמות מלאי",
    "משקל",
    "שפה",
    "שם קובץ תמונה",
    "מוצג",
    "מומלץ",
    "חדש",
    "פעיל",
]

sample = [
    "001",
    "729000000001",
    "שם הספר לדוגמה",
    "שבת",
    "ספרי יסוד|מומלצים",
    "שם המחבר",
    "שם הרב",
    "שם ההוצאה",
    "תיאור קצר של הספר",
    "תיאור ארוך יותר שיופיע בפרטי המוצר",
    120,
    99,
    25,
    0.8,
    "עברית",
    "001.jpg",
    "כן",
    "כן",
    "כן",
    "כן",
]

wb = Workbook()
wb.remove(wb.active)
wb.properties.title = "תבנית יבוא ספרים - אוצר הקדושה"
wb.properties.creator = "אוצר הקדושה"

walnut = "2A1B0E"
header_fill = PatternFill("solid", fgColor=walnut)
subheader_fill = PatternFill("solid", fgColor="FFF7D6")
required_fill = PatternFill("solid", fgColor="EAF7EA")
example_fill = PatternFill("solid", fgColor="F8FAFC")
thin = Side(style="thin", color="E5E7EB")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

instructions = wb.create_sheet("הוראות")
instructions.sheet_view.rightToLeft = True
instructions.freeze_panes = "A6"
instructions.column_dimensions["A"].width = 28
instructions.column_dimensions["B"].width = 90
instructions["A1"] = "תבנית יבוא ספרים - אוצר הקדושה"
instructions["A1"].font = Font(bold=True, size=18, color=walnut)
instructions["A2"] = "כל גיליון מייצג קטגוריה ראשית. מלאו את הספרים בכל גיליון ושמרו כקובץ Excel לפני היבוא."
instructions["A2"].font = Font(size=12, color="334155")

instruction_rows = [
    ("שלב 1", "מלאו ספרים בגיליונות הקטגוריות. אין לשנות את שמות העמודות באנגלית בשורה 1."),
    ("שלב 2", "צרפו תמונות לקובץ ZIP. תמונה ראשית לפי SKU: לדוגמה 001.jpg."),
    ("גלריה", "תמונות גלריה לפי SKU: לדוגמה 001-1.jpg, 001-2.jpg, 001-3.jpg."),
    ("קטגוריות נוספות", "בעמודת AdditionalCategories הפרידו קטגוריות באמצעות הסימן | לדוגמה: ספרי יסוד|מומלצים."),
    ("שדות חובה", "SKU, BookName, Price הם שדות חובה לכל ספר."),
    ("Boolean", "בעמודות Featured / Recommended / New / Active ניתן לרשום כן/לא או true/false."),
    ("עדכון מחירים", "למרכז עדכון מחירים מספיק קובץ עם SKU, Price, SalePrice."),
    ("עדכון מלאי", "למרכז עדכון מלאי מספיק קובץ עם SKU, StockQuantity."),
]
for row_index, (key, value) in enumerate(instruction_rows, start=4):
    instructions.cell(row_index, 1, key)
    instructions.cell(row_index, 2, value)
    for col in range(1, 3):
        cell = instructions.cell(row_index, col)
        cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        cell.border = border
        if col == 1:
            cell.fill = subheader_fill
            cell.font = Font(bold=True, color=walnut)

for category in categories:
    ws = wb.create_sheet(category)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2000"

    for col, header in enumerate(headers, start=1):
        c1 = ws.cell(1, col, header)
        c1.fill = header_fill
        c1.font = Font(bold=True, color="FFFFFF")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = border

        c2 = ws.cell(2, col, hebrew_labels[col - 1])
        c2.fill = required_fill if header in ["SKU", "BookName", "Price"] else subheader_fill
        c2.font = Font(bold=True, color=walnut)
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2.border = border

        c3 = ws.cell(3, col, sample[col - 1])
        c3.fill = example_fill
        c3.font = Font(color="64748B", italic=True)
        c3.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        c3.border = border

    widths = {
        "A": 14,
        "B": 18,
        "C": 34,
        "D": 22,
        "E": 34,
        "F": 24,
        "G": 24,
        "H": 24,
        "I": 42,
        "J": 55,
        "K": 12,
        "L": 12,
        "M": 14,
        "N": 10,
        "O": 14,
        "P": 22,
        "Q": 12,
        "R": 12,
        "S": 10,
        "T": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 48

    yes_no = DataValidation(type="list", formula1='"כן,לא,true,false"', allow_blank=True)
    ws.add_data_validation(yes_no)
    for col in ["Q", "R", "S", "T"]:
        yes_no.add(f"{col}3:{col}2000")

    lang = DataValidation(type="list", formula1='"עברית,אנגלית,יידיש,ארמית"', allow_blank=True)
    ws.add_data_validation(lang)
    lang.add("O3:O2000")

    for row in range(4, 2001):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            cell.border = border

price = wb.create_sheet("עדכון מחירים")
price.sheet_view.rightToLeft = True
for col, header in enumerate(["SKU", "Price", "SalePrice"], start=1):
    cell = price.cell(1, col, header)
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
    price.column_dimensions[get_column_letter(col)].width = 18
price.append(["001", 120, 99])

inventory = wb.create_sheet("עדכון מלאי")
inventory.sheet_view.rightToLeft = True
for col, header in enumerate(["SKU", "StockQuantity"], start=1):
    cell = inventory.cell(1, col, header)
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
    inventory.column_dimensions[get_column_letter(col)].width = 18
inventory.append(["001", 25])

wb.active = 0
wb.save(out_file)

check = load_workbook(out_file)
expected = ["הוראות"] + categories + ["עדכון מחירים", "עדכון מלאי"]
assert check.sheetnames == expected, check.sheetnames
for sheet_name in expected:
    assert check[sheet_name].sheet_view.rightToLeft is True

print(out_file.resolve())
